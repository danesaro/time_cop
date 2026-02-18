"""XLSX report generation service using openpyxl."""

import logging
from datetime import date
from decimal import Decimal
from io import BytesIO
from typing import Optional

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

logger = logging.getLogger(__name__)

# ── Styles ───────────────────────────────────────────────────────────────────

_HEADER_FONT = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
_HEADER_FILL = PatternFill(start_color="2B5797", end_color="2B5797", fill_type="solid")
_HEADER_ALIGNMENT = Alignment(horizontal="center", vertical="center", wrap_text=True)

_DATA_FONT = Font(name="Calibri", size=10)
_DATA_ALIGNMENT = Alignment(vertical="top", wrap_text=True)

_SUBTOTAL_FONT = Font(name="Calibri", size=10, bold=True)
_SUBTOTAL_FILL = PatternFill(start_color="D9E2F3", end_color="D9E2F3", fill_type="solid")

_TOTAL_FONT = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
_TOTAL_FILL = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")

_THIN_BORDER = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin"),
)

_COLUMNS = ["Fecha", "Proyecto", "Descripción", "Categoría", "Horas"]
_COL_WIDTHS = [14, 25, 50, 25, 10]

_CATEGORIA_LABELS = {
    "proyectoFacturable": "Proyecto Facturable",
    "proyectoNoFacturable": "Proyecto No Facturable",
    "otrosNoFacturable": "Otros No Facturable",
}


# ── Public API ───────────────────────────────────────────────────────────────

def generar_reporte_xlsx(
    registros: list[dict],
    mes: int,
    año: int,
    nombre_usuario: Optional[str] = None,
) -> BytesIO:
    """Generate a formatted XLSX report from time records.

    Args:
        registros: List of record dicts from the database.
        mes: Month number (1-12).
        año: Year.
        nombre_usuario: Optional user display name for the title.

    Returns:
        A BytesIO buffer containing the XLSX workbook.
    """
    meses_es = [
        "", "Enero", "Febrero", "Marzo", "Abril", "Mayo", "Junio",
        "Julio", "Agosto", "Septiembre", "Octubre", "Noviembre", "Diciembre",
    ]
    mes_nombre = meses_es[mes] if 1 <= mes <= 12 else str(mes)
    titulo = f"Reporte de Tiempos — {mes_nombre} {año}"
    if nombre_usuario:
        titulo += f" — {nombre_usuario}"

    wb = Workbook()
    ws = wb.active
    ws.title = f"{mes_nombre} {año}"

    # ── Title row ────────────────────────────────────────────────────────
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(_COLUMNS))
    title_cell = ws.cell(row=1, column=1, value=titulo)
    title_cell.font = Font(name="Calibri", size=14, bold=True, color="2B5797")
    title_cell.alignment = Alignment(horizontal="center")

    # ── Headers ──────────────────────────────────────────────────────────
    header_row = 3
    for col_idx, (header, width) in enumerate(zip(_COLUMNS, _COL_WIDTHS), 1):
        cell = ws.cell(row=header_row, column=col_idx, value=header)
        cell.font = _HEADER_FONT
        cell.fill = _HEADER_FILL
        cell.alignment = _HEADER_ALIGNMENT
        cell.border = _THIN_BORDER
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    # ── Data rows ────────────────────────────────────────────────────────
    current_row = header_row + 1

    # Sort by date
    registros_sorted = sorted(registros, key=lambda r: (r["fecha"], r["created_at"]))

    total_general = Decimal("0")
    totales_proyecto: dict[str, Decimal] = {}
    totales_categoria: dict[str, Decimal] = {}

    for reg in registros_sorted:
        fecha = reg["fecha"]
        if isinstance(fecha, date):
            fecha_str = fecha.strftime("%Y-%m-%d")
        else:
            fecha_str = str(fecha)

        horas = Decimal(str(reg["horas_estimadas"]))
        proyecto = reg["proyecto"]
        categoria = reg["categoria"]

        total_general += horas
        totales_proyecto[proyecto] = totales_proyecto.get(proyecto, Decimal("0")) + horas
        totales_categoria[categoria] = totales_categoria.get(categoria, Decimal("0")) + horas

        values = [fecha_str, proyecto, reg["descripcion"], _CATEGORIA_LABELS.get(categoria, categoria), float(horas)]
        for col_idx, val in enumerate(values, 1):
            cell = ws.cell(row=current_row, column=col_idx, value=val)
            cell.font = _DATA_FONT
            cell.alignment = _DATA_ALIGNMENT
            cell.border = _THIN_BORDER

        current_row += 1

    # ── Subtotals by project ─────────────────────────────────────────────
    current_row += 1
    ws.cell(row=current_row, column=1, value="Totales por Proyecto").font = Font(
        name="Calibri", size=11, bold=True, color="2B5797"
    )
    current_row += 1

    for proyecto, horas in sorted(totales_proyecto.items()):
        for col_idx in range(1, len(_COLUMNS) + 1):
            cell = ws.cell(row=current_row, column=col_idx)
            cell.fill = _SUBTOTAL_FILL
            cell.font = _SUBTOTAL_FONT
            cell.border = _THIN_BORDER
        ws.cell(row=current_row, column=2, value=proyecto)
        ws.cell(row=current_row, column=5, value=float(horas))
        current_row += 1

    # ── Subtotals by category ────────────────────────────────────────────
    current_row += 1
    ws.cell(row=current_row, column=1, value="Totales por Categoría").font = Font(
        name="Calibri", size=11, bold=True, color="2B5797"
    )
    current_row += 1

    for categoria, horas in sorted(totales_categoria.items()):
        for col_idx in range(1, len(_COLUMNS) + 1):
            cell = ws.cell(row=current_row, column=col_idx)
            cell.fill = _SUBTOTAL_FILL
            cell.font = _SUBTOTAL_FONT
            cell.border = _THIN_BORDER
        ws.cell(row=current_row, column=4, value=_CATEGORIA_LABELS.get(categoria, categoria))
        ws.cell(row=current_row, column=5, value=float(horas))
        current_row += 1

    # ── Grand total ──────────────────────────────────────────────────────
    current_row += 1
    for col_idx in range(1, len(_COLUMNS) + 1):
        cell = ws.cell(row=current_row, column=col_idx)
        cell.fill = _TOTAL_FILL
        cell.font = _TOTAL_FONT
        cell.border = _THIN_BORDER
    ws.cell(row=current_row, column=4, value="TOTAL GENERAL")
    ws.cell(row=current_row, column=5, value=float(total_general))

    # ── Save to buffer ───────────────────────────────────────────────────
    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    logger.info("Generated XLSX report: %s, %d records, %.1f total hours", titulo, len(registros), total_general)
    return buffer
