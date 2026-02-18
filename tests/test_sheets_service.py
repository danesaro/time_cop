"""Tests for the XLSX report generation service."""

from datetime import date, datetime
from decimal import Decimal
from io import BytesIO
from uuid import uuid4

import pytest
from openpyxl import load_workbook

from app.services.sheets_service import generar_reporte_xlsx


def _make_registro(
    fecha: date,
    proyecto: str,
    descripcion: str,
    categoria: str,
    horas: float,
) -> dict:
    """Factory for test record dicts."""
    return {
        "id": uuid4(),
        "fecha": fecha,
        "usuario_telegram_id": 123456,
        "descripcion": descripcion,
        "proyecto": proyecto,
        "categoria": categoria,
        "horas_estimadas": Decimal(str(horas)),
        "texto_original": "test",
        "created_at": datetime(2026, 2, 1, 9, 0),
        "updated_at": datetime(2026, 2, 1, 9, 0),
    }


class TestGenerarReporteXlsx:
    """Tests for XLSX report generation."""

    def test_generates_valid_xlsx(self):
        """Verify the output is a valid XLSX workbook."""
        registros = [
            _make_registro(date(2026, 2, 1), "Alpha", "Task A", "proyectoFacturable", 3.0),
            _make_registro(date(2026, 2, 1), "Beta", "Task B", "proyectoNoFacturable", 2.5),
            _make_registro(date(2026, 2, 2), "Alpha", "Task C", "proyectoFacturable", 4.0),
        ]

        buffer = generar_reporte_xlsx(registros, 2, 2026)
        assert isinstance(buffer, BytesIO)

        wb = load_workbook(buffer)
        ws = wb.active
        assert ws.title == "Febrero 2026"

    def test_data_rows_content(self):
        """Verify data rows contain correct values."""
        registros = [
            _make_registro(date(2026, 2, 5), "Proyecto X", "Reunión diaria", "proyectoFacturable", 1.0),
            _make_registro(date(2026, 2, 5), "Proyecto X", "Desarrollo feature", "proyectoFacturable", 5.0),
        ]

        buffer = generar_reporte_xlsx(registros, 2, 2026)
        wb = load_workbook(buffer)
        ws = wb.active

        # Header is row 3, data starts at row 4
        assert ws.cell(row=4, column=2).value == "Proyecto X"
        assert ws.cell(row=4, column=3).value == "Reunión diaria"
        assert ws.cell(row=5, column=5).value == 5.0

    def test_empty_records(self):
        """Verify report handles empty records gracefully."""
        buffer = generar_reporte_xlsx([], 1, 2026)
        wb = load_workbook(buffer)
        ws = wb.active
        assert ws.title == "Enero 2026"

    def test_with_user_name(self):
        """Verify user name appears in the title."""
        registros = [
            _make_registro(date(2026, 3, 1), "P1", "Task", "otrosNoFacturable", 2.0),
        ]

        buffer = generar_reporte_xlsx(registros, 3, 2026, nombre_usuario="Daniel")
        wb = load_workbook(buffer)
        ws = wb.active

        title_value = ws.cell(row=1, column=1).value
        assert "Daniel" in title_value
        assert "Marzo" in title_value

    def test_totals_present(self):
        """Verify project and category subtotals are generated."""
        registros = [
            _make_registro(date(2026, 2, 1), "Alpha", "T1", "proyectoFacturable", 3.0),
            _make_registro(date(2026, 2, 2), "Beta", "T2", "proyectoNoFacturable", 5.0),
        ]

        buffer = generar_reporte_xlsx(registros, 2, 2026)
        wb = load_workbook(buffer)
        ws = wb.active

        # Check that total hours sum is somewhere in the sheet
        all_values = []
        for row in ws.iter_rows(values_only=True):
            all_values.extend([v for v in row if v is not None])

        assert 8.0 in all_values  # Grand total = 3 + 5
        assert 3.0 in all_values  # Alpha subtotal
        assert 5.0 in all_values  # Beta subtotal
